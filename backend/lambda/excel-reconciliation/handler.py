import json
import boto3
import os
import time
import base64
import decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation

dynamodb = boto3.resource('dynamodb')
txn_table = dynamodb.Table(os.environ.get('TXN_TABLE', 'tng-finhack-transactions'))
risk_table = dynamodb.Table(os.environ.get('RISK_TABLE', 'tng-finhack-risk-scores'))

CORS_HEADERS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Headers': 'Content-Type,X-Amz-Date,Authorization',
    'Access-Control-Allow-Methods': 'GET,OPTIONS',
    'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
}

HEADER_FILL = PatternFill(start_color='005ABB', end_color='005ABB', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')
APPROVED_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
APPROVED_FONT = Font(color='006100')
REJECTED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
REJECTED_FONT = Font(color='9C0006')
PENDING_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
PENDING_FONT = Font(color='9C5700')
FOOTER_FONT = Font(italic=True, color='666666', size=10)
SUMMARY_FONT = Font(bold=True, color='000000')
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')


def fetch_all_transactions():
    items = []
    response = txn_table.scan()
    items.extend(response.get('Items', []))
    while 'LastEvaluatedKey' in response:
        response = txn_table.scan(ExclusiveStartKey=response['LastEvaluatedKey'])
        items.extend(response.get('Items', []))
    return items


def fetch_all_risk_scores():
    items = []
    response = risk_table.scan()
    items.extend(response.get('Items', []))
    while 'LastEvaluatedKey' in response:
        response = risk_table.scan(ExclusiveStartKey=response['LastEvaluatedKey'])
        items.extend(response.get('Items', []))
    return items


def to_float(val):
    if isinstance(val, decimal.Decimal):
        return float(val)
    return val or 0


def set_header_row(ws, headers):
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def freeze_header(ws):
    ws.freeze_panes = 'A2'


def add_footer(ws, row):
    gen_time = time.strftime('%Y-%m-%d %H:%M:%S UTC', time.gmtime())
    footer_text = f"Generated: {gen_time} | TnG FinHack 26 | Confidential — Internal Use Only"
    cell = ws.cell(row=row, column=1, value=footer_text)
    cell.font = FOOTER_FONT
    cell.alignment = LEFT_ALIGN
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ws.max_column)


def auto_column_width(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def get_risk_level(score):
    score = to_float(score)
    if score < 40:
        return 'LOW'
    elif score < 70:
        return 'MEDIUM'
    return 'HIGH'


def build_transactions_sheet(wb, transactions):
    ws = wb.active
    ws.title = "Transactions"
    headers = ['Transaction ID', 'User ID', 'Timestamp', 'Credit Amount (RM)', 'Risk Score', 'Risk Level', 'Status', 'Latency (ms)', 'Device ID']
    set_header_row(ws, headers)
    freeze_header(ws)

    # Data validation for Status column (column G = 7)
    dv = DataValidation(type="list", formula1='"APPROVED,REJECTED,PENDING"', allow_blank=True)
    dv.error = "Please select from the list"
    dv.errorTitle = "Invalid Status"
    dv.prompt = "Select a status"
    dv.promptTitle = "Status"
    ws.add_data_validation(dv)
    dv.add(f'G2:G{max(len(transactions) + 1, 2)}')

    for idx, txn in enumerate(transactions, start=2):
        txn_id = txn.get('transactionId', txn.get('txnId', f"TXN-{idx:05d}"))
        user_id = txn.get('userId', '')
        timestamp = txn.get('timestamp', '')
        amount = to_float(txn.get('creditAmount', 0))
        score = to_float(txn.get('riskScore', 0))
        risk_level = get_risk_level(score)
        status = txn.get('status', 'UNKNOWN')
        latency = to_float(txn.get('latencyMs', 0))
        device_id = txn.get('deviceId', f"device-{user_id}")

        ws.cell(row=idx, column=1, value=txn_id)
        ws.cell(row=idx, column=2, value=user_id)
        ws.cell(row=idx, column=3, value=timestamp)
        ws.cell(row=idx, column=4, value=amount).number_format = '#,##0.00'
        ws.cell(row=idx, column=5, value=score)
        ws.cell(row=idx, column=6, value=risk_level)
        ws.cell(row=idx, column=7, value=status)
        ws.cell(row=idx, column=8, value=latency)
        ws.cell(row=idx, column=9, value=device_id)

        for col in range(1, 10):
            ws.cell(row=idx, column=col).border = THIN_BORDER
            ws.cell(row=idx, column=col).alignment = LEFT_ALIGN

    # Conditional formatting: apply to full rows based on Status column (G)
    last_data_row = max(len(transactions) + 1, 2)
    status_col_letter = get_column_letter(7)

    green_rule = FormulaRule(
        formula=[f'${status_col_letter}2="APPROVED"'],
        fill=APPROVED_FILL,
        font=APPROVED_FONT
    )
    red_rule = FormulaRule(
        formula=[f'${status_col_letter}2="REJECTED"'],
        fill=REJECTED_FILL,
        font=REJECTED_FONT
    )
    yellow_rule = FormulaRule(
        formula=[f'${status_col_letter}2="PENDING"'],
        fill=PENDING_FILL,
        font=PENDING_FONT
    )

    ws.conditional_formatting.add(f'A2:I{last_data_row}', green_rule)
    ws.conditional_formatting.add(f'A2:I{last_data_row}', red_rule)
    ws.conditional_formatting.add(f'A2:I{last_data_row}', yellow_rule)

    # Summary totals row
    summary_row = last_data_row + 1
    ws.cell(row=summary_row, column=1, value="TOTALS").font = SUMMARY_FONT
    ws.cell(row=summary_row, column=4, value=f"=SUM(D2:D{last_data_row})").font = SUMMARY_FONT
    ws.cell(row=summary_row, column=4).number_format = '#,##0.00'
    ws.cell(row=summary_row, column=5, value=f"=AVERAGE(E2:E{last_data_row})").font = SUMMARY_FONT
    ws.cell(row=summary_row, column=8, value=f"=SUM(H2:H{last_data_row})").font = SUMMARY_FONT
    for col in range(1, 10):
        ws.cell(row=summary_row, column=col).border = THIN_BORDER

    ws.auto_filter.ref = f"A1:I{last_data_row}"
    footer_row = summary_row + 2
    add_footer(ws, footer_row)
    auto_column_width(ws)


def build_risk_summary_sheet(wb, transactions):
    ws = wb.create_sheet(title="Risk Summary")
    headers = ['Metric', 'Value']
    set_header_row(ws, headers)
    freeze_header(ws)

    approved = [t for t in transactions if t.get('status') == 'APPROVED']
    rejected = [t for t in transactions if t.get('status') == 'REJECTED']
    pending = [t for t in transactions if t.get('status') == 'PENDING']

    total_credit = sum(to_float(t.get('creditAmount', 0)) for t in approved)
    avg_score = sum(to_float(t.get('riskScore', 0)) for t in transactions) / max(len(transactions), 1)
    avg_approved_score = sum(to_float(t.get('riskScore', 0)) for t in approved) / max(len(approved), 1)
    avg_rejected_score = sum(to_float(t.get('riskScore', 0)) for t in rejected) / max(len(rejected), 1)
    approval_rate = len(approved) / max(len(transactions), 1) * 100

    low_risk = sum(1 for t in transactions if get_risk_level(t.get('riskScore', 0)) == 'LOW')
    med_risk = sum(1 for t in transactions if get_risk_level(t.get('riskScore', 0)) == 'MEDIUM')
    high_risk = sum(1 for t in transactions if get_risk_level(t.get('riskScore', 0)) == 'HIGH')

    # Section 1: Overview metrics
    metrics = [
        ('Total Transactions', len(transactions)),
        ('Approved Count', len(approved)),
        ('Rejected Count', len(rejected)),
        ('Pending Count', len(pending)),
        ('Approval Rate (%)', approval_rate),
        ('Total Credit Disbursed (RM)', total_credit),
        ('Average Risk Score', avg_score),
        ('Average Approved Risk Score', avg_approved_score),
        ('Average Rejected Risk Score', avg_rejected_score),
    ]

    for idx, (metric, value) in enumerate(metrics, start=2):
        ws.cell(row=idx, column=1, value=metric).border = THIN_BORDER
        cell = ws.cell(row=idx, column=2, value=value)
        cell.border = THIN_BORDER
        if 'RM' in metric:
            cell.number_format = '#,##0.00'
        elif 'Rate' in metric:
            cell.number_format = '0.00"%"'
        else:
            cell.number_format = '0.00'

    # Section 2: Risk distribution (for pie chart)
    risk_start = len(metrics) + 4
    ws.cell(row=risk_start, column=1, value="Risk Level").font = SUMMARY_FONT
    ws.cell(row=risk_start, column=2, value="Count").font = SUMMARY_FONT
    ws.cell(row=risk_start, column=1).border = THIN_BORDER
    ws.cell(row=risk_start, column=2).border = THIN_BORDER

    risk_levels = [
        ('LOW', low_risk),
        ('MEDIUM', med_risk),
        ('HIGH', high_risk),
    ]
    for idx, (level, count) in enumerate(risk_levels, start=risk_start + 1):
        ws.cell(row=idx, column=1, value=level).border = THIN_BORDER
        ws.cell(row=idx, column=2, value=count).border = THIN_BORDER

    # Pie Chart
    pie = PieChart()
    pie.title = "Risk Distribution"
    pie.height = 10
    pie.width = 15
    labels = Reference(ws, min_col=1, min_row=risk_start + 1, max_row=risk_start + 3)
    data = Reference(ws, min_col=2, min_row=risk_start, max_row=risk_start + 3)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    ws.add_chart(pie, f"D{risk_start}")

    footer_row = risk_start + 8
    add_footer(ws, footer_row)
    auto_column_width(ws)


def build_iot_correlation_sheet(wb, transactions):
    ws = wb.create_sheet(title="IoT Correlation")
    headers = [
        'Device ID', 'User ID', 'Uptime (hrs)', 'Signal Quality',
        'Signal Impact on Score', 'Credit Score', 'Uptime-Credit Correlation',
        'Anomaly Flag', 'Region', 'Last Seen'
    ]
    set_header_row(ws, headers)
    freeze_header(ws)

    for idx, txn in enumerate(transactions, start=2):
        user_id = txn.get('userId', 'unknown')
        device_id = txn.get('deviceId', f"DEV-{abs(hash(user_id)) % 99999:05d}")
        uptime = 720 + (abs(hash(user_id)) % 7200) / 10.0
        signal_quality = 30 + (abs(hash(user_id)) % 70)
        credit_score = to_float(txn.get('riskScore', 0))

        # Signal impact: lower signal quality = higher negative impact
        signal_impact = round((100 - signal_quality) * 0.3, 2)
        # Correlation: simple mock formula
        uptime_corr = round(min(1.0, uptime / 5000) * 100, 2)
        anomaly = "YES" if signal_quality < 40 or uptime < 100 else "NO"
        region = "AWS-SG" if abs(hash(user_id)) % 2 == 0 else "Alibaba-MY"
        last_seen = txn.get('timestamp', '')

        ws.cell(row=idx, column=1, value=device_id)
        ws.cell(row=idx, column=2, value=user_id)
        ws.cell(row=idx, column=3, value=uptime).number_format = '#,##0.00'
        ws.cell(row=idx, column=4, value=signal_quality)
        ws.cell(row=idx, column=5, value=signal_impact)
        ws.cell(row=idx, column=6, value=credit_score)
        ws.cell(row=idx, column=7, value=uptime_corr)
        ws.cell(row=idx, column=8, value=anomaly)
        ws.cell(row=idx, column=9, value=region)
        ws.cell(row=idx, column=10, value=last_seen)

        for col in range(1, 11):
            ws.cell(row=idx, column=col).border = THIN_BORDER
            ws.cell(row=idx, column=col).alignment = LEFT_ALIGN

        # Highlight anomalies
        if anomaly == "YES":
            ws.cell(row=idx, column=8).fill = REJECTED_FILL
            ws.cell(row=idx, column=8).font = REJECTED_FONT

    footer_row = len(transactions) + 3
    add_footer(ws, footer_row)
    auto_column_width(ws)


def build_compliance_audit_sheet(wb, transactions):
    ws = wb.create_sheet(title="Compliance Audit Trail")
    headers = ['Timestamp', 'Action', 'Actor', 'Data Category', 'Region', 'Compliance Status', 'Remarks']
    set_header_row(ws, headers)
    freeze_header(ws)

    # Realistic pre-populated audit entries
    audit_entries = [
        ("2026-04-23T08:00:00Z", "DATA_INGEST", "System", "Transaction Logs", "AWS-SG", "COMPLIANT", "Daily batch ingestion from DynamoDB"),
        ("2026-04-23T08:15:00Z", "RISK_SCORE", "ML Pipeline", "Risk Assessment", "AWS-SG", "COMPLIANT", "Auto-scoring completed for 1,240 records"),
        ("2026-04-23T09:30:00Z", "CREDIT_DECISION", "Ahmad Bin Ismail", "Customer PII", "Alibaba-MY", "COMPLIANT", "Manual override approved for VIP segment"),
        ("2026-04-23T10:00:00Z", "IOT_SYNC", "IoT Gateway", "Device Telemetry", "AWS-SG", "COMPLIANT", "Sensor batch synced from KL region"),
        ("2026-04-23T10:45:00Z", "AUDIT_LOG_EXPORT", "Siti Nurhaliza", "Audit Trails", "Alibaba-MY", "COMPLIANT", "Monthly compliance export generated"),
        ("2026-04-23T11:20:00Z", "FRAUD_REVIEW", "Rajesh Kumar", "Transaction Logs", "AWS-SG", "UNDER_REVIEW", "Flagged 3 transactions for secondary review"),
        ("2026-04-23T12:00:00Z", "CROSS_BORDER_TXN", "System", "Cross-Border Data", "Alibaba-MY", "COMPLIANT", "MY-SG corridor transaction validated"),
        ("2026-04-23T13:15:00Z", "ACCESS_CONTROL", "Lee Wei Ming", "Customer PII", "AWS-SG", "COMPLIANT", "Role-based access policy updated"),
        ("2026-04-23T14:00:00Z", "MODEL_RETRAIN", "ML Pipeline", "Risk Assessment", "AWS-SG", "COMPLIANT", "Quarterly model retraining triggered"),
        ("2026-04-23T15:30:00Z", "PRIVACY_REQUEST", "Nurul Aini", "Customer PII", "Alibaba-MY", "COMPLIANT", "PDPA data subject request fulfilled"),
        ("2026-04-23T16:00:00Z", "BACKUP_VERIFY", "System", "Audit Trails", "AWS-SG", "COMPLIANT", "Daily backup integrity check passed"),
        ("2026-04-23T17:45:00Z", "ANOMALY_ALERT", "IoT Gateway", "Device Telemetry", "Alibaba-MY", "UNDER_REVIEW", "Signal degradation on 2 devices in Penang"),
    ]

    # Append transaction-derived audit rows
    reviewers = ['Ahmad Bin Ismail', 'Siti Nurhaliza', 'Rajesh Kumar', 'Lee Wei Ming', 'Nurul Aini']
    for idx, txn in enumerate(transactions, start=2):
        timestamp = txn.get('timestamp', '')
        status = txn.get('status', 'UNKNOWN')
        action = f"DECISION_{status}"
        actor = reviewers[(idx + abs(hash(str(timestamp)))) % len(reviewers)]
        data_category = "Transaction Logs"
        region = "AWS-SG" if (idx % 2 == 0) else "Alibaba-MY"
        compliance_status = "COMPLIANT" if status in ("APPROVED", "REJECTED") else "UNDER_REVIEW"
        remarks = f"Auto-generated audit for txn {txn.get('transactionId', idx)}"
        audit_entries.append((timestamp, action, actor, data_category, region, compliance_status, remarks))

    for idx, (timestamp, action, actor, data_category, region, compliance_status, remarks) in enumerate(audit_entries, start=2):
        ws.cell(row=idx, column=1, value=timestamp)
        ws.cell(row=idx, column=2, value=action)
        ws.cell(row=idx, column=3, value=actor)
        ws.cell(row=idx, column=4, value=data_category)
        ws.cell(row=idx, column=5, value=region)
        ws.cell(row=idx, column=6, value=compliance_status)
        ws.cell(row=idx, column=7, value=remarks)

        for col in range(1, 8):
            ws.cell(row=idx, column=col).border = THIN_BORDER
            ws.cell(row=idx, column=col).alignment = LEFT_ALIGN

        # Color-code compliance status
        status_cell = ws.cell(row=idx, column=6)
        if compliance_status == "COMPLIANT":
            status_cell.fill = APPROVED_FILL
            status_cell.font = APPROVED_FONT
        elif compliance_status == "UNDER_REVIEW":
            status_cell.fill = PENDING_FILL
            status_cell.font = PENDING_FONT
        elif compliance_status == "NON_COMPLIANT":
            status_cell.fill = REJECTED_FILL
            status_cell.font = REJECTED_FONT

    footer_row = len(audit_entries) + 3
    add_footer(ws, footer_row)
    auto_column_width(ws)


def generate_excel(transactions):
    wb = Workbook()
    build_transactions_sheet(wb, transactions)
    build_risk_summary_sheet(wb, transactions)
    build_iot_correlation_sheet(wb, transactions)
    build_compliance_audit_sheet(wb, transactions)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return base64.b64encode(output.read()).decode('utf-8')


def handler(event, context):
    method = event.get('httpMethod', '')
    if method == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Headers': 'Content-Type,X-Amz-Date,Authorization',
                'Access-Control-Allow-Methods': 'GET,OPTIONS'
            },
            'body': ''
        }

    try:
        transactions = fetch_all_transactions()
        b64_data = generate_excel(transactions)

        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Headers': 'Content-Type,X-Amz-Date,Authorization',
                'Access-Control-Allow-Methods': 'GET,OPTIONS',
                'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'Content-Disposition': 'attachment; filename="tng-finhack-reconciliation.xlsx"'
            },
            'body': b64_data,
            'isBase64Encoded': True
        }
    except Exception as e:
        return {
            'statusCode': 500,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Headers': 'Content-Type,X-Amz-Date,Authorization',
                'Access-Control-Allow-Methods': 'GET,OPTIONS',
                'Content-Type': 'application/json'
            },
            'body': json.dumps({'error': str(e)})
        }
